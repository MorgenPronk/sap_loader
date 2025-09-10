# %%
print("Importing libraries...")
import sys
import pandas as pd
import os
from openpyxl import load_workbook
from typing import NamedTuple, Any, Union, List, Dict
import logging
import json
import shutil
from tqdm import tqdm

# %%
def config_logging(filepath):

    filepath = os.path.abspath(filepath)

    if not os.path.exists(filepath):
        with open(filepath, 'w') as f:
            f.write("")

    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        filename=filepath,
        filemode='a',
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

# %%
def load_files(equipment_path, hierarchy_path, loadsheet_path):
    equip_df = pd.read_excel(equipment_path)
    hierarchy_df = pd.read_excel(hierarchy_path)
    loadsheet_dict = pd.read_excel(loadsheet_path, sheet_name= None)
    return equip_df, hierarchy_df, loadsheet_dict

# %%
def preprocess_dataframes(hierarchy_df, equip_df):
    # Pre processing of some of the dataframes

    # hierarchy_df
    # Create normalized tags for comparison later
    # Normalize Columns
    for col in ['level_6_1', 'level_7', 'level_8']:
        if col in hierarchy_df.columns:
            hierarchy_df[f'{col}_normalized'] = hierarchy_df[col].astype(str).str.replace('-', '').str.strip()
        else:
            logging.warning(f"Column '{col}' not found in the Excel file.")

    # equip_df
    # Replace any spaces with "_" in the column
    # This is important for the Named tuples we use later
    equip_df.columns = equip_df.columns.str.replace(" ", "_")

# %%
def static_variables():
    #  Defining some of the output columns and rows for the simpleload.xlsx file
    # We will define maps here for LLM outputs to code if we need to, but will likely just train LLM to output the right code.
    

    # extract needed values from JSON config file
    # we will hardcode the config file here for now
    json_path = "./data/config.json"
    with open(json_path, 'r') as f:
        config = json.load(f)
    row_start = config.get("row_start")

    # columns for "FLOC Only" sheet
    FLOC_sheet = config.get("FLOC_sheet")

    # columns for FLOCEquip Sheet
    equip_sheet = config.get("equip_sheet")
    
    return row_start, FLOC_sheet, equip_sheet

def get_hierarchy_desc(normalized_id):
    # print(normalized_id)
    hierarchy_desc = None
    path = os.path.join("data", "config.json")
    with open(path, 'r') as f:
        data = json.load(f) # this is the map for the descriptions we are just using the names in the config file for the target keys
    target_keys = ['L4_codes', 'L4_1_codes', 'L5_codes', 'L5_1_codes']
    
    for target_key in target_keys:
        subdict = data[target_key]
        # print(f"{target_key}") # debugging
        for k,v in subdict.items():
            # print(f" {v}") debugging
            if str(v).replace('-', '') == normalized_id:
                hierarchy_desc = k
                break
    # print(f"  {hierarchy_desc}") # debugging
    return hierarchy_desc

# %%
def get_hierarchy_chain(start_id, hierarchy_df, equip_df) -> List[Dict[str, Any]]:
    normalized_id = start_id.replace('-', '').strip()
    # print(f"Normalized id: {normalized_id}")
    columns = ['level_4', 'level_4_1', 'level_5', 'level_5_1', 'level_6', 'level_6_1', 'level_7', 'level_8']
    normalized_columns = ['level_4', 'level_4_1', 'level_5', 'level_5_1', 'level_6', 'level_6_1_normalized', 'level_7_normalized', 'level_8_normalized']
    equipment_levels = ['level_6_1', 'level_7', 'level_8']
    levels_to_concat = ['level_4', 'level_4_1', 'level_5', 'level_5_1', 'level_6']

    chain = []
    match_copy = None

    while normalized_id:
        # Find the matching row in the hierarchy_df
        match_row = None
        match_col = None

        for norm_col in normalized_columns:
            print(f"Nomalized_id: {normalized_id}; norm_col: {norm_col}") # Debugging
            match = hierarchy_df[hierarchy_df[norm_col] == normalized_id]
            print(f"match: {match}") #Debugging
            if not match.empty:
                match_row = match.iloc[0].copy()
                # print(f"~match_row when created~\n {match_row}") # Debugging 
                match_col = norm_col.replace('_normalized', '')
                # print(f"match_row: {match_row}; match_col: {match_col}") # Debugging 
                break

        if match_row is None:
            logging.warning(f"No match found in hierarchy for ID '{normalized_id}'")
            # We need to check if it is an equipment level - If it is, even if it isn't in the hierarchy, we need to grab the right information
            if match_col not in equipment_levels:
                break

        # Now we have a match_row and match_col, we should build out what the tags are in the chain and put them in a list
        # we take the row, and start at the beginnging of the row, and then we will build out the chain

        # copy the match to get an uncatentated version - the hierarchy output is the just the category the equipment will be in. We have to make the tags for each of those which involves concatenating things
        concat_values = match_row[levels_to_concat].copy() # This isn't concatenated yet
        match_row_non_concat = match_row.copy()

        # Here we concatenate the hierarchy levels with the prior levels to get the tag for any given level
        for i, level in enumerate(levels_to_concat):
            match_row[level] = '-'.join(
                str(concat_values[l]) for l in levels_to_concat[:i+1] if pd.notna(match_row[l])
            )


        print(f"~match_row before extraction~\n {match_row}")
        current_id = match_row[match_col]
        idx = columns.index(match_col)
        parent_id = match_row[columns[idx - 1]] if idx > 0 else None
        parent_id_not_concat = match_row_non_concat[columns[idx-1]] if idx > 0 else None # this is to get the non concatenated version to keep the loop going
        subclass = match_row.get('subclass', '')
        print(f"current_id- {current_id}; idx- {idx}; parnet_id- {parent_id}; subclass- {subclass} ")

        entry = {
            'ID': current_id,
            'Superior FLOC': parent_id,
        }

        # If it's an equipment level, try to get extra data
        if match_col in equipment_levels:
            
            subclass = match_row.get('subclass', '')

            equip_match = equip_df[
                (equip_df['Serial_Number'].astype(str).str.strip() == current_id) |
                (equip_df['Tag_Number'].astype(str).str.strip() == current_id)
            ]
            # print(f"equip_match: {equip_match}")

            if not equip_match.empty:
                equip_row = equip_match.iloc[0]
                entry.update({
                    'Subclass': match_row.get('subclass', ''),
                    'Make': str(equip_row['Mfg_Desc']).strip(),
                    'Model': str(equip_row['Product_Model']).strip(),
                    'Description': "; ".join(
                        str(equip_row[col]).strip()
                        for col in ['Description', 'Description_2', 'Description_3']
                        if pd.notna(equip_row[col])
                    )
                })
                
                # print(f"entry_update: row-{equip_row} ; {entry}")
            else:
                logging.warning(f"Equipment data not found for ID '{current_id}'")
        else:
            # If it's not equipment level, we will look for a match and give it a description

            hierarchy_desc = get_hierarchy_desc(normalized_id)
            if hierarchy_desc is not None:
                entry.update({
                    'Description': hierarchy_desc
                })

        chain.append(entry)

        # now we change the id to the level below and do everything again
        normalized_id = str(parent_id_not_concat).replace('-', '').strip() if parent_id else None

    print(f"chain:\n {chain}")
    return chain

# %%
def write_chain_to_output(ws, chain, row_start, FLOC_sheet, current_row_offset=0):
    for i, entry in enumerate(chain):
        ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["ID (Blank if Equipment)"], value=entry['ID'])
        ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["Superior FLOC (Parent)"], value=entry['Superior FLOC'])
        ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["Class (DCAM Subclass)"], value=entry.get('Subclass', ''))

        if 'Description' in entry:
            ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["Description"], value=entry['Description'])
        if 'Make' in entry:
            ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["Make"], value=entry['Make'])
        if 'Model' in entry:
            ws.cell(row=row_start + current_row_offset + i, column=FLOC_sheet["Model"], value=entry['Model'])

    current_row_offset += len(chain)

    return current_row_offset, ws

# %%
def extract_from_sheets(equip_df, hierarchy_df, row_start, loadsheet_path, FLOC_sheet):

    wb = load_workbook(loadsheet_path)
    ws = wb[FLOC_sheet["sheet_name"]] # This will need to be put in the for loop, when we need to move between multiple sheets - like if we need to fill out the FLOCEquip sheet 
    current_row_offset = 0
    # print(equip_df) # debugging
    # input() # debugging
    for row in tqdm(equip_df.itertuples(index=True), total=len(equip_df), desc= "Processing Equipment"): # Take head() out when you want to run the entire sheet
        # print("in progress bar loop")
        serial = str(row.Serial_Number).strip()
        tag = str(row.Tag_Number).strip()

        # Get the serial Number
        if serial and tag:
            id = serial
        elif serial:
            id = serial
        elif tag:
            id = tag
        else:
            id = ''

        # get the hierarchy chain for the current ID
        chain = get_hierarchy_chain(id, hierarchy_df, equip_df)
        
        current_row_offset, ws = write_chain_to_output(ws, chain, row_start, FLOC_sheet, current_row_offset)

        # print(f"writing to {loadsheet_path}. Chain: {chain}") # Debugging
        wb.save(loadsheet_path)

def excel_row_to_df_index_equip(excel_row):
    # Calibration to help start from the endpoint. The input is the excel row number where the next equipment is that should go into the output should be.
    return excel_row - 2

# %%
def main():
    # Define the files that we are working with
    log_output_path = 'equipment_match.log'
    equipment_path = 'data/Current JDE Equipment Table_NoFilter.xlsx' #'./data/Current JDE Equipment Table 6-2-25.xlsx'
    hierarchy_path = './data/hierarchy_output.xlsx'
    loadsheet_path = './data/simpleload.xlsx'

    # Copy and rename a new loadsheet - This keeps a blank copy of the loadsheet for future use that doesn't get touched
    
    new_loadsheet_path = loadsheet_path.replace('.xlsx', '_output.xlsx')
    if os.path.exists(new_loadsheet_path):
        print("Output file exists. Will resume from last written row")
        equipment_start_index = excel_row_to_df_index_equip(10540)
        output_start_row = 59936 # excel row
    else:
        equipment_start_index = 0
        output_start_row = 3
        shutil.copy2(loadsheet_path, new_loadsheet_path)

    loadsheet_path = new_loadsheet_path

    # Configure the logging
    config_logging(log_output_path)

    # Load files into Dataframes and dictionaries for use using pandas
    equip_df, hierarchy_df, loadsheet_dict = load_files(equipment_path, hierarchy_path, loadsheet_path)

    # Preprocess dataframes - create normailized equipment columns with normalized tags for the hierarchy df and Replace any spaces with "_" in the column headers for the equipment df
    preprocess_dataframes(hierarchy_df, equip_df)

    # Define the start of rows in the FLOC_sheet, Define the columns in the FLOC and equipment sheet
    # equip_sheet is the excel sheet that hast the equipment from JDE listed on it, compes from equipment_path
    # FLOC_sheet is one of the sheets 'FLOC Only' sheet from the loadsheet and comes from loadsheet path. This is our output.
    row_start, FLOC_sheet, equip_sheet = static_variables()

    # If we want to start from a different row then we overwrite the row_start variable
    row_start = output_start_row

    equip_df = equip_df.head(1).copy() # For testing purposes, we will only take the first 5 rows of the hierarchy_df. Remove this line when you want to run the entire hierarchy_df

    # run the sheet
    extract_from_sheets(equip_df.iloc[equipment_start_index:].copy(), hierarchy_df, row_start, loadsheet_path, FLOC_sheet)
    
if __name__ == "__main__":
    print("Chromes burning...")
    main()
    print("Ran hot, cooled clean")

